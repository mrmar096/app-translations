from lxml import etree as ET
from openpyxl import load_workbook
import os
import json

def load_android_translations():
    """
    Load translations from the Android Excel sheet.
    """
    texts_android = {}
    android_language_mapper = {
        'SPANISH': 'values',
        'ENGLISH': 'values-en',
        'FRENCH': 'values-fr'
        # Adjust keys as needed
    }

    # Initialize dictionary for Android changes
    changes_android = {}

    # Load Android workbook
    wb_android = load_workbook('app_translations.xlsx')
    sheet = wb_android['Android']  # Get the sheet named "Android"

    # Print values from translations in the Excel sheet
    for row in range(2, sheet.max_row + 1):
        key = sheet.cell(row=row, column=1).value
        translations = {}
        for col in range(2, sheet.max_column + 1):
            android_lang = android_language_mapper.get(sheet.cell(row=1, column=col).value, None)
            if android_lang is not None:
                translation = sheet.cell(row=row, column=col).value
                translations[android_lang] = translation
        texts_android[key] = translations

    return texts_android, android_language_mapper, changes_android


def create_update_xml_file(android_lang, texts_android, changes_android):
    """
    Create or update XML file for Android translations.
    """
    xml_file = os.path.join(android_lang, 'strings.xml')

    if not os.path.exists(xml_file):
        # Create a new XML file and write translations from the Excel sheet
        root = ET.Element('resources')
        for key, value in texts_android.items():
            if android_lang in value and value[android_lang] != 'UNTRANSLATED':
                string_elem = ET.SubElement(root, 'string', name=key)
                string_elem.text = value[android_lang]

        tree = ET.ElementTree(root)
        with open(xml_file, 'wb') as file:
            tree.write(file, encoding='utf-8', xml_declaration=True, pretty_print=True)

        print(f"\nCreated and populated file: {xml_file}")
    else:
        # Update existing XML file with new translations
        print(f"\nFile: {xml_file}")
        tree = ET.parse(xml_file)
        root_elem = tree.getroot()
        # Initialize changes_android[android_lang] if not exists
        if android_lang not in changes_android:
            changes_android[android_lang] = {}
        # Read existing xml keys
        existing_keys = {string_elem.get('name') for string_elem in root_elem.findall('.//string')}

        for key, value in texts_android.items():
            if key not in existing_keys:
                # If the key does not exist, add a new line with the translation from the Excel sheet
                excel_value = value[android_lang].strip() if value[android_lang] else None
                if excel_value is not None:
                    if android_lang in value and excel_value != 'UNTRANSLATED':
                        new_string_elem = ET.Element('string', name=key)
                        new_string_elem.text = excel_value
                        root_elem.append(new_string_elem)
                        changes_android[android_lang][key] = excel_value
                        print(f"Added new key \"{key}\" to {android_lang} with value {excel_value}")

        if changes_android[android_lang]:
            # Apply indentation
            ET.indent(tree, space="\t", level=0)
            with open(xml_file, 'w', encoding='utf-8') as file:
                file.write(ET.tostring(tree, encoding='utf-8', xml_declaration=True, pretty_print=True).decode())

        for string_elem in root_elem.findall('.//string'):
            key = string_elem.get('name')
            if key in texts_android and android_lang in texts_android[key]:
                xml_value = string_elem.text.strip() if string_elem.text else None
                excel_value = texts_android[key][android_lang].strip() if texts_android[key][android_lang] else None
                if excel_value is not None:
                    if excel_value != 'UNTRANSLATED' and (xml_value is None or excel_value != xml_value):
                        print(f"Updated {key} in {android_lang} to {excel_value}")
                        changes_android[android_lang][key] = excel_value
                        string_elem.text = excel_value

        if changes_android[android_lang]:
            # Apply indentation
            ET.indent(tree, space="\t", level=0)
            with open(xml_file, 'w', encoding='utf-8') as file:
                file.write(ET.tostring(tree, encoding='utf-8', xml_declaration=True, pretty_print=True).decode())

        return changes_android



# Main script
texts_android, android_language_mapper, changes_android = load_android_translations()

print("Processing localization files for Android:")

for android_lang in android_language_mapper.values():
    if not os.path.exists(android_lang):
        os.makedirs(android_lang)

    changes_android = create_update_xml_file(android_lang, texts_android, changes_android)

# Print information about the changes
print("\nChanges:")
if all(not changes for changes in changes_android.values()):
    print("No changes")
else:
    print(json.dumps(changes_android, indent=2, ensure_ascii=False).encode('utf-8').decode())

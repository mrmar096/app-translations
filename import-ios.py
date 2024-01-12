import os
from openpyxl import load_workbook
import codecs
import json

def load_ios_translations():
    """
    Load translations from the iOS Excel sheet.
    """
    texts_ios = {}
    ios_language_mapper = {
        'SPANISH': 'es.lproj',
        'ENGLISH': 'en.lproj',
        'FRENCH': 'fr.lproj'
        # Add more mappings as needed
    }

    # Initialize dictionary for iOS changes
    changes_ios = {}

    # Load iOS workbook
    wb_combined = load_workbook('app_translations.xlsx')
    sheet_ios = wb_combined['iOS']  # Get the sheet named "iOS"

    # Print values from translations in the Excel sheet
    for row in range(2, sheet_ios.max_row + 1):
        key = sheet_ios.cell(row=row, column=1).value
        translations = {}
        for col in range(2, sheet_ios.max_column + 1):
            ios_lang = ios_language_mapper.get(sheet_ios.cell(row=1, column=col).value, None)
            if ios_lang is not None:
                translation = sheet_ios.cell(row=row, column=col).value
                translations[ios_lang] = translation
        texts_ios[key] = translations

    return texts_ios, ios_language_mapper, changes_ios


def create_update_strings_file(ios_lang, texts_ios, changes_ios):
    """
    Create or update .strings file for iOS translations.
    """
    strings_file = os.path.join(ios_lang, 'Localizable.strings')

    if not os.path.exists(strings_file):
        # Create a new .strings file and write translations from the Excel sheet
        with open(strings_file, 'w', encoding='utf-8') as ios_file:
            for key, value in texts_ios.items():
                if ios_lang in value and value[ios_lang] != 'UNTRANSLATED':
                    ios_file.write("\"{}\" = \"{}\";\n".format(key, value[ios_lang]))
        print(f"\nCreated and populated file: {strings_file}")
    else:
        # Update existing .strings file with new translations
        print(f"\nFile: {strings_file}")
        with open(strings_file, 'r', encoding='utf-8') as ios_file:
            content = ios_file.readlines()

        # Initialize changes_android[android_lang] if not exists
        if ios_lang not in changes_ios:
            changes_ios[ios_lang] = {}

        # Create a set of existing keys in the .strings file
        existing_keys = {line.split('=')[0].strip()[1:-1] for line in content if '=' in line}

        with open(strings_file, 'a', encoding='utf-8') as ios_file:
            for key, value in texts_ios.items():
                # Check if the key already exists in the .strings file
                if key not in existing_keys:
                    # If the key does not exist, add a new line with the translation from the Excel sheet
                    excel_value = value[ios_lang].strip() if value[ios_lang] else None
                    if excel_value is not None:
                        if ios_lang in value and excel_value != 'UNTRANSLATED':
                            ios_file.write("\"{}\" = \"{}\";\n".format(key, excel_value))
                            changes_ios[ios_lang][key] = excel_value
                            print(f"Added new key \"{key}\" to {ios_lang} with value {excel_value}")
            ios_file.close()

        # Continue with updating translations in an existing .strings file
        with open(strings_file, 'r', encoding='utf-8') as ios_file:
            content = ios_file.readlines()

        for i in range(len(content)):
            line = content[i]
            parts = line.split("=")
            if len(parts) == 2:
                key = parts[0].strip()[1:-1]

                if key in texts_ios and ios_lang in texts_ios[key]:
                    strings_value = parts[1].strip()[1:-2].strip() if parts[1].strip() else None
                    excel_value = texts_ios[key][ios_lang].strip() if texts_ios[key][ios_lang] else None
                    if excel_value is not None:
                        if strings_value is None or (excel_value != strings_value and excel_value != 'UNTRANSLATED'):
                            print(f"Updating {key} in {ios_lang} to {excel_value}")
                            content[i] = "\"{}\" = \"{}\";\n".format(key, excel_value)

                            # Record the change
                            if ios_lang not in changes_ios:
                                changes_ios[ios_lang] = {}
                            if key not in changes_ios[ios_lang]:
                                changes_ios[ios_lang][key] = excel_value

        if len(changes_ios) != 0:
            output = ''.join(content)
            with codecs.open(strings_file, 'w', 'utf-8') as ios_file:
                ios_file.write(output)
                ios_file.close()

    return changes_ios


# Main script for iOS
texts_ios, ios_language_mapper, changes_ios = load_ios_translations()

print("Processing localization files for iOS:")

for ios_lang in ios_language_mapper.values():
    if not os.path.exists(ios_lang):
        os.makedirs(ios_lang)

    changes_ios = create_update_strings_file(ios_lang, texts_ios, changes_ios)

# Print information about the changes
print("Changes:")
if all(not changes for changes in changes_ios.values()):
    print("No Changes")
else:
    print(json.dumps(changes_ios, indent=2, ensure_ascii=False).encode('utf-8').decode())
